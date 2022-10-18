#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

total_list = []
col_grade = dict()
total_sorted = []
A = []
B = []
B2 = []
C = []
i = 0

row_id = 1;
for row in ws:
    if row_id != 1:
        midterm = ws.cell(row = row_id, column = 3).value
        final= ws.cell(row = row_id, column = 4).value
        homework = ws.cell(row = row_id, column = 5).value
        attendance = ws.cell(row = row_id, column = 6).value
        total = midterm * 0.3 + final * 0.35 + homework * 0.34 + attendance
        col_grade[row_id] = total
        ws.cell (row = row_id, column = 7, value = total)
    row_id += 1

#grade

total_sorted = sorted(col_grade.items(), key = lambda x : x[1], reverse = True)

for col in range(len(total_sorted)):
    if col <= len(total_sorted) * 0.3 - 1:
        A.append(total_sorted[col])
    else:
        B.append(total_sorted[col])
    if total_sorted[col][1] < 40 :
        ws.cell(row = total_sorted[col][0], column = 8, value = 'F')

for col in range(len(A)):
    if col <= len(A) * 0.5 - 1:
        ws.cell(row = A[col][0], column = 8, value = 'A+')
    else:
        ws.cell(row = A[col][0], column = 8, value = 'A0')
    if A[col][1] < 40:
        ws.cell(row = A[col][0], column = 8, value = 'F')

for col in range(len(B)):
    if col <= len(B) * (4 / 7) - 1:
        B2.append(B[col])
    else:
        C.append(B[col])
    if B[col][1] < 40:
        ws.cell(row = B[col][0], column = 8, value = 'F')

for col in range(len(B2)):
    if col <= len(B2) * 0.5 - 1:
        ws.cell(row = B2[col][0], column = 8, value = 'B+')
    else:
        ws.cell(row = B2[col][0], column = 8, value = 'B0')
    if B2[col][1] < 40:
        ws.cell(row = B2[col][0], column = 8, value = 'F')

for col in range(len(C)):
    if col <= len(C) * 0.5 - 1:
        ws.cell(row = C[col][0], column = 8, value = 'C+')
    else:
        ws.cell(row = C[col][0], column = 8, value = 'C0')
    if C[col][1] < 40 :
        ws.cell(row = C[col][0], column = 8, value = 'F')

wb.save( "student.xlsx" )

#print(total_sort)
#print(total_sorted)

